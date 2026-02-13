#!/usr/bin/env python3
"""
Build a Zilliz prospecting spreadsheet from structured prospect data.

Usage:
    python scripts/build_spreadsheet.py <input_json> <output_xlsx>

Input JSON format:
{
    "metadata": {
        "industry": "legaltech",
        "region": "APAC",
        "date": "2026-02-13",
        "pain_point": "...",
        "differentiator": "...",
        "region_focus": "...",
        "social_proof": "..."
    },
    "prospects": [
        {
            "rank": 1,
            "tier": "Tier 1",
            "company": "Company Name",
            "hq": "City, Country",
            "funding": "Public (TICKER, $XB mkt cap)",
            "product_focus": "Description...",
            "why_zilliz_fit": "Argument...",
            "current_infra": "OpenSearch (confirmed via AWS case study)",
            "evidence_strength": "HIGH - AWS published case study",
            "key_contacts": "Name (Title)\\nLinkedIn URL\\nName2 (Title2)",
            "pitch_angle": "Specific pitch...",
            "lead_warmth_score": 9,
            "scoring_breakdown": {
                "evidence_strength": 9,
                "vector_fit": 8,
                "scale_pain": 8,
                "reachability": 7,
                "strategic_value": 8
            }
        }
    ],
    "scoring_overrides": null
}
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Color constants
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
TIER1_FILL = PatternFill('solid', fgColor='E2EFDA')
TIER2_FILL = PatternFill('solid', fgColor='FFF2CC')
TIER3_FILL = PatternFill('solid', fgColor='F2F2F2')
MATRIX_HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
SUMMARY_HEADER_FILL = PatternFill('solid', fgColor='1F4E79')

BODY_FONT = Font(size=10, name='Arial')
BOLD_FONT = Font(bold=True, size=10, name='Arial')
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')
TOP_ALIGN = Alignment(vertical='top')
CENTER_ALIGN = Alignment(horizontal='center', vertical='top')
THIN_BORDER = Border(
    bottom=Side(style='thin', color='D9D9D9')
)

COLUMN_WIDTHS = {'A': 6, 'B': 8, 'C': 25, 'D': 20, 'E': 30,
                 'F': 50, 'G': 55, 'H': 45, 'I': 30, 'J': 45,
                 'K': 55, 'L': 12}

HEADERS = ['Rank', 'Tier', 'Company', 'HQ', 'Funding',
           'Product / Focus', 'Why Zilliz is a Good Fit',
           'Likely Vector DB / Search Infra', 'Evidence Strength',
           'Key Contacts to Reach Out', 'Entry Point / Pitch Angle',
           'Lead Warmth Score (1-10)']

WRAP_COLS = {6, 7, 8, 10, 11}  # F, G, H, J, K (1-indexed)


def get_tier_fill(tier: str) -> PatternFill:
    if 'Tier 1' in tier:
        return TIER1_FILL
    elif 'Tier 2' in tier:
        return TIER2_FILL
    return TIER3_FILL


def build_prospect_tab(wb: Workbook, prospects: list) -> None:
    ws = wb.active
    ws.title = 'Prospect List'

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    ws.row_dimensions[1].height = 35
    ws.auto_filter.ref = f'A1:L{len(prospects) + 1}'
    ws.freeze_panes = 'A2'

    # Data rows
    for row_idx, p in enumerate(prospects, 2):
        values = [
            p['rank'], p['tier'], p['company'], p['hq'],
            p['funding'], p['product_focus'], p['why_zilliz_fit'],
            p['current_infra'], p['evidence_strength'],
            p['key_contacts'], p['pitch_angle'], p['lead_warmth_score']
        ]
        tier_fill = get_tier_fill(p['tier'])

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.fill = tier_fill
            cell.border = THIN_BORDER
            if col_idx in WRAP_COLS:
                cell.alignment = WRAP_ALIGN
            elif col_idx in (1, 12):
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = TOP_ALIGN

        ws.row_dimensions[row_idx].height = 120


def build_scoring_matrix_tab(wb: Workbook, overrides=None) -> None:
    ws = wb.create_sheet('Scoring Matrix')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50

    # Title
    ws['A1'] = 'Lead Warmth Scoring Matrix'
    ws['A1'].font = Font(bold=True, size=14, name='Arial')
    ws.merge_cells('A1:E1')

    ws['A3'] = 'Formula: (E × 0.30) + (V × 0.25) + (S × 0.20) + (R × 0.15) + (SV × 0.10)'
    ws['A3'].font = Font(italic=True, size=10, name='Arial')
    ws.merge_cells('A3:E3')

    # Headers
    matrix_headers = ['Dimension', 'Weight', 'Low (1-3)', 'Medium (4-6)', 'High (7-10)']
    for col_idx, header in enumerate(matrix_headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.fill = MATRIX_HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    weights = overrides if overrides else {
        'evidence_strength': 30,
        'vector_fit': 25,
        'scale_pain': 20,
        'reachability': 15,
        'strategic_value': 10
    }

    dimensions = [
        {
            'name': 'Evidence Strength',
            'weight': f"{weights['evidence_strength']}%",
            'low': 'No public info about tech stack. Inference based on industry patterns only. No job postings, case studies, or tech blog mentions.',
            'medium': 'Indirect signals: job postings mention relevant tech, known cloud provider usage, or industry peers have published architecture details.',
            'high': 'Direct confirmed evidence: published case study naming technologies, engineer conference talk, tech blog post, or job posting explicitly mentioning vector search scaling.'
        },
        {
            'name': 'Vector Search / GenAI Fit',
            'weight': f"{weights['vector_fit']}%",
            'low': 'Vector search is tangential to core product. Primary use case is structured data or workflows. AI features are marketing-level.',
            'medium': 'Clear use cases for vector search exist but not the primary product driver. Building AI capabilities where vector search fits.',
            'high': 'Vector search is central to core product. Search, matching, recommendation are fundamental. Semantic search is a competitive imperative.'
        },
        {
            'name': 'Scale & Pain Probability',
            'weight': f"{weights['scale_pain']}%",
            'low': 'Small-medium data volumes. Under 1M documents, low QPS. Current infra likely sufficient. Cost/performance benefits of purpose-built vector DB are marginal.',
            'medium': 'Meaningful scale: millions of documents, thousands QPS, or growing volumes that will stress infra within 12-18 months.',
            'high': 'Large scale: 100M+ documents/vectors, high QPS, real-time updates. OpenSearch/ES kNN hits known limits: memory-heavy HNSW, 3-replica costs, segment merge latency.'
        },
        {
            'name': 'Reachability & Timing',
            'weight': f"{weights['reachability']}%",
            'low': 'No identifiable decision-makers. Large bureaucratic org with no visible AI initiative. No recent triggers. Cold outreach only.',
            'medium': 'Decision-makers identifiable on LinkedIn but no warm connection or timing trigger. Stable company, no recent changes.',
            'high': 'Clear decision-makers with public profiles. Recent trigger: new funding, AI launch, infra migration, new CTO hire, or engineer quoted in case study.'
        },
        {
            'name': 'Strategic Value',
            'weight': f"{weights['strategic_value']}%",
            'low': 'Small or unknown company. Limited signaling value. Win not referenceable for other prospects.',
            'medium': 'Recognized company in their industry. Useful for vertical-specific sales. Moderate brand recognition.',
            'high': 'Marquee logo recognizable to every prospect in region/industry. Win unlocks doors. Strong case study and marketing potential.'
        }
    ]

    for row_offset, dim in enumerate(dimensions):
        row = 6 + row_offset
        ws.cell(row=row, column=1, value=dim['name']).font = BOLD_FONT
        ws.cell(row=row, column=2, value=dim['weight']).font = BODY_FONT
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        for col_idx, key in enumerate(['low', 'medium', 'high'], 3):
            cell = ws.cell(row=row, column=col_idx, value=dim[key])
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGN
        ws.row_dimensions[row].height = 80


def build_summary_tab(wb: Workbook, prospects: list, metadata: dict) -> None:
    ws = wb.create_sheet('Summary')

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 80

    title = f"{metadata.get('industry', '')} {metadata.get('region', '')} Prospect Summary"
    ws['A1'] = title.strip()
    ws['A1'].font = Font(bold=True, size=14, name='Arial')
    ws.merge_cells('A1:B1')

    tier_counts = {'Tier 1': 0, 'Tier 2': 0, 'Tier 3': 0}
    evidence_counts = {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
    scores = []

    for p in prospects:
        tier = p['tier']
        if tier in tier_counts:
            tier_counts[tier] += 1
        ev = p.get('evidence_strength', '').split(' - ')[0].strip().upper()
        if ev in evidence_counts:
            evidence_counts[ev] += 1
        if isinstance(p.get('lead_warmth_score'), (int, float)):
            scores.append(p['lead_warmth_score'])

    avg_score = round(sum(scores) / len(scores), 1) if scores else 0

    stats = [
        ('Total Prospects', len(prospects)),
        ('Tier 1 (Confirmed Evidence + Strong Fit)', tier_counts['Tier 1']),
        ('Tier 2 (Medium Evidence + Good Fit)', tier_counts['Tier 2']),
        ('Tier 3 (Inferred Evidence + Emerging Fit)', tier_counts['Tier 3']),
        ('', ''),
        ('Average Lead Warmth Score', avg_score),
        ('Prospects with HIGH Evidence', evidence_counts['HIGH']),
        ('Prospects with MEDIUM Evidence', evidence_counts['MEDIUM']),
        ('Prospects with LOW Evidence', evidence_counts['LOW']),
    ]

    row = 3
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = BODY_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='Key Positioning').font = Font(bold=True, size=12, name='Arial')
    row += 1

    positioning = [
        ('Primary Pain Point', metadata.get('pain_point', '')),
        ('Zilliz Differentiator', metadata.get('differentiator', '')),
        ('Region Focus', metadata.get('region_focus', '')),
        ('Social Proof', metadata.get('social_proof', '')),
    ]

    for label, value in positioning:
        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = BODY_FONT
        cell.alignment = WRAP_ALIGN
        ws.row_dimensions[row].height = 40
        row += 1


def main():
    if len(sys.argv) < 3:
        print("Usage: python build_spreadsheet.py <input.json> <output.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(input_path) as f:
        data = json.load(f)

    wb = Workbook()

    prospects = data.get('prospects', [])
    metadata = data.get('metadata', {})
    scoring_overrides = data.get('scoring_overrides', None)

    build_prospect_tab(wb, prospects)
    build_scoring_matrix_tab(wb, scoring_overrides)
    build_summary_tab(wb, prospects, metadata)

    wb.save(output_path)
    print(f"Spreadsheet saved to: {output_path}")


if __name__ == '__main__':
    main()
